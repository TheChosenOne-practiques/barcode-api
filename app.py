from flask import Flask, request, send_file, abort
from barcode import Code128
from barcode.writer import ImageWriter
import io
import os
import openpyxl

app = Flask(__name__)

if not os.path.exists("barcodes"):
    os.makedirs("barcodes")

excel_path = "barcodes/registros.xlsx"

if not os.path.exists(excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Nº"
    ws["B1"] = "Código"
    wb.save(excel_path)

@app.route("/")
def home():
    return '''
<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Generador de Códigos de Barras</title>
  <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@400;500&display=swap" rel="stylesheet">
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    body {
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      background: #0f0f0f;
      font-family: "IBM Plex Sans", sans-serif;
    }

    .card {
      background: #1a1a1a;
      border: 1px solid #2e2e2e;
      border-radius: 12px;
      padding: 40px 36px;
      width: 100%;
      max-width: 420px;
      box-shadow: 0 8px 32px rgba(0,0,0,0.4);
    }

    .logo {
      font-family: "IBM Plex Mono", monospace;
      font-size: 11px;
      letter-spacing: 3px;
      text-transform: uppercase;
      color: #555;
      margin-bottom: 28px;
    }

    h1 {
      font-size: 22px;
      font-weight: 500;
      color: #f0f0f0;
      margin-bottom: 8px;
    }

    p {
      font-size: 13px;
      color: #666;
      margin-bottom: 28px;
      line-height: 1.5;
    }

    label {
      display: block;
      font-size: 11px;
      letter-spacing: 1.5px;
      text-transform: uppercase;
      color: #555;
      margin-bottom: 8px;
      font-family: "IBM Plex Mono", monospace;
    }

    input {
      width: 100%;
      padding: 12px 14px;
      background: #111;
      border: 1px solid #2e2e2e;
      border-radius: 8px;
      color: #f0f0f0;
      font-family: "IBM Plex Mono", monospace;
      font-size: 15px;
      outline: none;
      transition: border-color 0.2s;
      margin-bottom: 16px;
    }

    input:focus { border-color: #e8c84a; }
    input::placeholder { color: #3a3a3a; }

    button {
      width: 100%;
      padding: 13px;
      background: #e8c84a;
      color: #0f0f0f;
      border: none;
      border-radius: 8px;
      font-family: "IBM Plex Mono", monospace;
      font-size: 13px;
      font-weight: 600;
      letter-spacing: 1px;
      cursor: pointer;
      transition: opacity 0.2s, transform 0.1s;
    }

    button:hover { opacity: 0.88; }
    button:active { transform: scale(0.98); }
    button:disabled { opacity: 0.4; cursor: not-allowed; }

    .status {
      margin-top: 14px;
      font-size: 12px;
      font-family: "IBM Plex Mono", monospace;
      color: #555;
      min-height: 18px;
      text-align: center;
    }
  </style>
</head>
<body>
  <div class="card">
    <div class="logo">Barcode API</div>
    <h1>Generador de códigos</h1>
    <p>Introduce un número para generar y descargar su código de barras.</p>

    <label for="codigo">Número</label>
    <input type="text" id="codigo" placeholder="Ej: 1234567890" autocomplete="off">
    <button id="btn" onclick="generar()">Generar y descargar</button>
    <div class="status" id="status"></div>
  </div>

  <script>
    async function generar() {
      const valor = document.getElementById("codigo").value.trim();
      const btn = document.getElementById("btn");
      const status = document.getElementById("status");

      if (!valor) {
        status.textContent = "⚠ Introduce un número primero.";
        return;
      }

      btn.disabled = true;
      btn.textContent = "Generando...";
      status.textContent = "";

      // Primero generamos la imagen en el servidor
      const res = await fetch("/generate?data=" + encodeURIComponent(valor));

      if (!res.ok) {
        status.textContent = "✗ Error al generar el código.";
        btn.disabled = false;
        btn.textContent = "Generar y descargar";
        return;
      }

      const urlImg = "/img/barcode_" + valor + ".png";

      // Abrir en nueva pestaña
      window.open(urlImg, "_blank");

      // Descargar
      const link = document.createElement("a");
      link.href = urlImg;
      link.download = "barcode_" + valor + ".png";
      document.body.appendChild(link);
      link.click();
      document.body.removeChild(link);

      btn.disabled = false;
      btn.textContent = "Generar y descargar";
      status.textContent = "✓ Código generado para: " + valor;
    }

    document.getElementById("codigo").addEventListener("keydown", e => {
      if (e.key === "Enter") generar();
    });
  </script>
</body>
</html>
    '''

@app.route("/generate")
def generate():
    data = request.args.get("data")
    if not data:
        return "Falta data", 400

    buffer = io.BytesIO()
    codigo = Code128(data, writer=ImageWriter())
    codigo.write(buffer)
    buffer.seek(0)

    filename = f"barcode_{data}.png"
    filepath = f"barcodes/{filename}"

    with open(filepath, "wb") as f:
        f.write(buffer.getvalue())

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    next_row = ws.max_row + 1
    ws[f"A{next_row}"] = next_row - 1
    ws[f"B{next_row}"] = data
    wb.save(excel_path)

    return "OK"

@app.route("/img/<filename>")
def get_image(filename):
    filepath = f"barcodes/{filename}"
    if os.path.exists(filepath):
        return send_file(filepath, mimetype="image/png")
    else:
        abort(404)

if __name__ == "__main__":
    app.run()
